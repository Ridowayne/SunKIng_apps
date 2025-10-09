import re
import time
import anthropic
from django.shortcuts import redirect, render 
import PyPDF2
import requests
from django.http import HttpResponse, JsonResponse
from .models import Category, Question, Answer, Customer, CallAudit
import random
import json
import os
from django.core.files.storage import default_storage
from django.conf import settings
from openai import OpenAI
from dotenv import load_dotenv
# import os
import time
import openai
import pytesseract
from pdf2image import convert_from_path
import pdfplumber
import base64
import anthropic
import tempfile
import logging
import boto3
from botocore.exceptions import ClientError
from pathlib import Path
import google.generativeai as genai

load_dotenv()


# Create your views here.

def get_all_questions(request):
    if request.method == 'GET':
        category_data = []
        country = request.user.country if request.user.is_authenticated else 'Nigeria'
        all_categories = Category.objects.filter(archived=False, country=country)
        
        for category in all_categories:
            # Get all questions for the current category
            all_questions = list(Question.objects.filter(category=category, archived=False))
            
            
            random.shuffle(all_questions)
            selected_questions = all_questions[:2]

            question_data = []
            for  question in selected_questions:                
                answer_data = []
                # Get all answers for the current question
                answers = Answer.objects.filter(question=question)
                for answer in answers:
                    answer_data.append({
                        'answer': answer.answer,
                        'id': answer.uid,  
                    })
                
                question_data.append({
                    'question': question.question,  
                    'answers': answer_data,
                    "id": question.uid, 
                })
            
            category_data.append({
                'category': category.category_name, 
                'category_number': category.category_number,
                'questions': question_data,
            })
       
            

        return render(request, 'customer_test_questions.html', {'category_data': category_data})
    elif request.method == 'POST':
        # get all the answers that the user selected and find them, then add all the points
        user_score = 0
        data = request.POST

        for question_id in data:
            if question_id.startswith('question_'):
                user_answer = data.get(question_id)

                if user_answer:
                    try:
                        # Fetch the answer object by its unique ID
                        answer = Answer.objects.get(uid=user_answer)
                        user_score += answer.marks

                    except Answer.DoesNotExist:
                        print(f'Answer with uid {user_answer} not found')
                        continue

        # Create the new customer record after processing all answers
        new_customer_record = Customer(
            name=request.POST.get('customer_name'),
            total_score=user_score,
            product_category=request.POST.get('product_category'),
            area=request.POST.get('area'),
            country = request.user.country if request.user.is_authenticated else 'Nigeria'
        )
        new_customer_record.save()

        # Redirect after saving the customer
        return redirect('customer')

    return JsonResponse({'error': 'No responses received', 'score': user_score})

        
def category_functions(request):
    if request.method == 'GET':
        country = request.user.country if request.user.is_authenticated else 'Nigeria'
        cats = Category.objects.filter(archived=False, country=country).order_by('-created_at')
        cats_data = []
        for cat in cats:
            cats_data.append({
                'name': cat.category_name,
                'category_id': cat.uid,
            })
        return render(request, 'category.html', {'categories': cats_data})
    elif request.method == 'POST':
        country = request.user.country if request.user.is_authenticated else 'Nigeria'
        available_cats = Category.objects.filter(archived=False, country=country).order_by('-created_at')
        count = available_cats.count()
        new_category = Category(
           category_name= request.POST.get('category_name'),
            category_number = count+1,
            country = request.user.get('country'),
       )
        new_category.save()
        return HttpResponse(new_category)
    elif request.method == 'PUT':
        try:
            
            data = json.loads(request.body)
            category_uid = data.get('category')  

            # Fetch the question instance
            category = Category.objects.get(uid=category_uid)
            category.archived = True  
            category.save() 

            return JsonResponse({'status': 'success', 'message': 'category archived successfully.'}, status=200)
        
        except Category.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Category not found.'}, status=404)
        


def questions_functions(request):
    if request.method == 'GET':
        country = request.user.country if request.user.is_authenticated else 'Nigeria'
        cats = Category.objects.filter(archived=False, country=country).order_by('-created_at')
        cats_data = []
        for cat in cats:
            cats_data.append({
                'name': cat.category_name,
                'category_id': cat.uid,
            })
        questions = []
        ques = Question.objects.select_related('category').filter(archived=False) 
        for que in ques:
            
            questions.append({
                'question': que.question,
                'category': que.category.category_name,
                'question_id': que.uid,  
            })
           
            
        return render(request, 'questions.html', {'questions': questions, 'cats': cats_data})
    elif request.method == 'POST':
       
        new_question = Question(
            category_id=request.POST.get('category'),  
            question=request.POST.get('question'),
        )
        new_question.save()
        answers = request.POST.getlist('answers')  
        marks = request.POST.getlist('marks')     
        
        # Ensure we have 4 answers and marks
        if len(answers) == 4 and len(marks) == 4:
            for answer, mark in zip(answers, marks):
                new_answer = Answer(
                    question=new_question,
                    answer=answer, 
                    marks=mark,     
                )
                new_answer.save()
            return JsonResponse({'status': 'your question has been saved', 'question': new_question.question})
        else:
            return JsonResponse({'error': 'You must provide exactly 4 answers and marks.'}, status=400)
    elif request.method == 'PUT':
        try:
            
            data = json.loads(request.body)
            question_uid = data.get('question')  

            # Fetch the question instance
            question = Question.objects.get(uid=question_uid)
            question.archived = True  
            question.save() 

            return JsonResponse({'status': 'success', 'message': 'Question archived successfully.'}, status=200)
        
        except Question.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Question not found.'}, status=404)
        

def customer_functions(request):
    if request.method == 'GET':
        customer_data = []
        # I should add filtering by country after I get clarifications on it
        country = request.user.country if request.user.is_authenticated else 'Nigeria'
        customers = Customer.objects.filter(country=country).order_by('-created_at')
       
        for customer in customers:
            customer_rating = ''
            if customer.total_score >= 32:
                customer_rating = 'Excelent'
            elif customer.total_score >= 24 and customer.total_score <= 31:
                customer_rating = 'Good'
            elif customer.total_score >=16 and customer.total_score <= 23:
                customer_rating = 'Moderate'
            else:
                customer_rating = 'Poor'
            customer_data.append({
                'name': customer.name,
                'product': customer.product_category,
                'area': customer.area,
                'country': customer.country,
                'score': customer.total_score,
                'ratings': customer_rating,
            })
            # print(customer_data)
        return render(request, 'customers.html', {'customers': customer_data})


def extract_text_from_pdf(pdf_path):
    """Extract text from text-based PDFs using pdfplumber"""
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text += page.extract_text() + "\n"
    return text


def call_gpt(prompt):
   
    api_key = os.environ.get('OPENAI_API_KEY', os.getenv("OPENAI_API_KEY"))
    url = 'https://api.openai.com/v1/chat/completions'

    try:
        payload = {
            'model': 'gpt-3.5-turbo',
            'messages': [
                {'role': 'system', 'content': 'You are a helpful assistant.'},
                {'role': 'user', 'content': f'Accurately extract the financial statement this which will be used for further credit analysis: {prompt}'}
            ]
        }
        
        headers = {
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {api_key}'
        }

        response = requests.post(url, data=json.dumps(payload), headers=headers)
        response.raise_for_status()  # Raises exception for error``
        data = response.json()
        
        return data

    except requests.exceptions.RequestException as e:
        print(f'Request error: {e}')
    except json.JSONDecodeError as e:
        print(f'Error parsing JSON response: {e}')


# Load environment variables once at startup
load_dotenv()

# Initialize OpenAI client once
client = OpenAI(api_key =os.getenv("OPENAI_API_KEY"))

def call_gpt_assistant(request):
    if request.method == 'GET':
        return render(request, 'analysis.html')
    
    elif request.method == 'POST':
        # Check if file was uploaded
        if 'file' not in request.FILES:
            return render(request, 'analysis.html', {'error': 'No file uploaded'})
        
        uploaded_file = request.FILES['file']
        
        try:
            # Save the file to media/uploads/ directory
            file_path = default_storage.save(f"uploads/{uploaded_file.name}", uploaded_file)
            absolute_file_path = os.path.join(settings.MEDIA_ROOT, file_path)

            # Process PDF with OpenAI
            analysis_result = process_pdf_with_openai(absolute_file_path)
            print(analysis_result)
            
            return render(request, 'analysis.html', {'analysis': analysis_result})
        
        except Exception as e:
            print(e)
            return render(request, 'analysis.html', {'error': str(e)})

def process_pdf_with_openai(pdf_path):
    try:
        # Upload PDF file to OpenAI
        with open(pdf_path, 'rb') as pdf_file:
            uploaded_file = client.files.create(
                file=pdf_file,
                purpose="assistants"
            )

        # Create thread with message and file attachment
        thread = client.beta.threads.create(
            messages=[
                {
                    "role": "user",
                    "content": "Analyze this PDF and extract the financial statements, The password to the document is 101997. Focus on balance sheets, income statements, and cash flow statements.",
                    "attachments": [  # Changed to 'attachments' parameter
                        {"file_id": uploaded_file.id, "tools": [{"type": "code_interpreter"}]}
                    ]
                }
            ]
        )

        # Create assistant with retrieval capability
        assistant = client.beta.assistants.create(
            name="Financial Analyst GPT",
            instructions="You are an expert financial analyst. Analyze the provided financial documents and extract key metrics, trends, and insights.",
            tools=[
                {"type": "code_interpreter"},
                {"type": "file_search"}  
            ],
            model="gpt-3.5-turbo"
        )

        # Create and monitor the run
        run = client.beta.threads.runs.create(
            thread_id=thread.id,
            assistant_id=assistant.id
        )

        # Poll for completion
        while run.status not in ['completed', 'failed', 'cancelled']:
            time.sleep(7)
            run = client.beta.threads.runs.retrieve(
                thread_id=thread.id,
                run_id=run.id
            )

        if run.status == 'completed':
            print("conpleted")
            messages = client.beta.threads.messages.list(
                thread_id=thread.id,
                order="asc"
            )
            # Extract assistant's response
            for msg in reversed(messages.data):
                if msg.role == 'assistant' and msg.content:
                    return msg.content[0].text.value
            
            return "No analysis found in response"
        elif run.status == 'failed':
            print("call failed")
            return "Analysis failed"
        else:
        
            return f"Analysis failed with status: {run.status}"

    except Exception as e:
        print(e)
        raise Exception(f"OpenAI processing error: {str(e)}")

def call_claude_assistant(request):
    if request.method == 'GET':
        return render(request, 'analysis.html')
    
    elif request.method == 'POST':
        # Check if file was uploaded
        if 'file' not in request.FILES:
            return render(request, 'analysis.html', {'error': 'No file uploaded'})
        
        uploaded_file = request.FILES['file']
        
        try:
            # Save the file to media/uploads/ directory
            file_path = default_storage.save(f"uploads/{uploaded_file.name}", uploaded_file)
            absolute_file_path = os.path.join(settings.MEDIA_ROOT, file_path)

            # Process PDF with Claude
            analysis_result = process_pdf_with_gpt(absolute_file_path)
            
            return render(request, 'analysis.html', {'analysis': analysis_result})
        
        except Exception as e:
            print(e)
            return render(request, 'analysis.html', {'error': str(e)})

def call_claude_assistant(request):
    if request.method == 'GET':
        return render(request, 'analysis.html')
    
    elif request.method == 'POST':
        # Check if file was uploaded
        if 'file' not in request.FILES:
            return render(request, 'analysis.html', {'error': 'No file uploaded'})
        
        uploaded_file = request.FILES['file']
        
        try:
            # Save the file to media/uploads/ directory
            file_path = default_storage.save(f"uploads/{uploaded_file.name}", uploaded_file)
            absolute_file_path = os.path.join(settings.MEDIA_ROOT, file_path)

            # Process PDF with Claude
            analysis_result = process_pdf_with_gpt(absolute_file_path)
            
            return render(request, 'analysis.html', {'analysis': analysis_result})
        
        except Exception as e:
            print(e)
            return render(request, 'analysis.html', {'error': str(e)})
        
os.getenv("OPENAI_API_KEY")
def call_claude_assistant(request):
    if request.method == 'GET':
        return render(request, 'analysis.html')
    
    elif request.method == 'POST':
        # Check if file was uploaded
        if 'file' not in request.FILES:
            return render(request, 'analysis.html', {'error': 'No file uploaded'})
        
        uploaded_file = request.FILES['file']
        
        try:
            # Save the file to media/uploads/ directory
            file_path = default_storage.save(f"uploads/{uploaded_file.name}", uploaded_file)
            absolute_file_path = os.path.join(settings.MEDIA_ROOT, file_path)

            # Process PDF with Claude
            analysis_result = process_pdf_with_gpt(absolute_file_path)
            
            # Clean up the uploaded file
            cleanup_uploaded_file(absolute_file_path)
            
            return render(request, 'analysis.html', {'analysis': analysis_result})
        
        except Exception as e:
            print(e)
            return render(request, 'analysis.html', {'error': str(e)})

def process_pdf_with_claude(pdf_path):
    try:
        # Initialize Claude client
        client = anthropic.Client(
            api_key=os.getenv("ANTHROPIC_API_KEY")
        )
        
        # Read and encode PDF file
        with open(pdf_path, 'rb') as pdf_file:
            pdf_base64 = base64.b64encode(pdf_file.read()).decode('utf-8')
        
        # Create message with the PDF content
        message = client.messages.create(
            model="claude-2",
            max_tokens=4096,
            messages=[{
                "role": "user",
                "content": "Analyze this PDF and extract the financial statements. "
                          "The password to the document is 101997. "
                          "Focus on balance sheets, income statements, and cash flow statements. "
                          f"Here is the base64 encoded PDF content: {pdf_base64}"
            }]
        )

        # Extract and return the analysis
        if hasattr(message, 'content') and message.content:
            return message.content[0].text
        else:
            return "No analysis found in response"

    except Exception as e:
        print(e)
        raise Exception(f"Claude processing error: {str(e)}")

def cleanup_uploaded_file(file_path):
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except Exception as e:
        print(f"Error cleaning up file: {e}")

            
        


# OpenAI API Key
openai.api_key = os.getenv("OPENAI_API_KEY")


# Function to extract text from scanned PDFs (image-based)
def extract_text_from_scanned_pdf(pdf_path):
    # Convert PDF pages to images
    images = convert_from_path(pdf_path)
    text = ""
    for img in images:
        # Use Tesseract OCR to extract text from the images
        text += pytesseract.image_to_string(img) + "\n"
    return text

def process_pdf_with_gpt(text):
    client = openai.OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    
    # with open(pdf_path, 'rb') as pdf_file:
    #     pdf_base64 = base64.b64encode(pdf_file.read()).decode('utf-8')
    
    response = client.chat.completions.create(
        model="gpt-3.5-turbo",
        messages=[
            {"role": "system", "content": "You are a financial analyst."},
            {"role": "user", "content": f"acuurately extract and arrange accounting data from this document a: {text}"}
        ]
    )
    
    return response.choices[0].message.content



# Django view function to handle PDF uploads and process them
def upload_pdf(request):
    """Handle PDF uploads and process them with OpenAI."""
    if request.method == 'GET':
        # Render the initial upload page with the form
        return render(request, 'analysis.html')

    elif request.method == 'POST':
        if 'file' not in request.FILES:
            # Return an error if no file is uploaded
            return render(request, 'analysis.html', {'error': 'No file uploaded'})

        uploaded_file = request.FILES['file']

        # Save the file to media/uploads/ directory
        try:
            file_path = default_storage.save(f"uploads/{uploaded_file.name}", uploaded_file)
            absolute_file_path = os.path.join(settings.MEDIA_ROOT, file_path)
        except Exception as e:
            # Handle any errors during file saving
            return render(request, 'analysis.html', {'error': f'Error saving file: {str(e)}'})

        # Initialize text variable to hold extracted text
        text = ""

        try:
            # First, try extracting text from the PDF as a text-based document
            text = extract_text_from_pdf(absolute_file_path)

            # If the text extraction is empty (indicating it's likely a scanned PDF)
            if not text:
                raise ValueError("No text extracted from PDF.")
        except Exception as e:
            # If no text is extracted or an error occurs, treat it as a scanned (image-based) PDF
            # text = extract_text_from_scanned_pdf(absolute_file_path)
            # if not text:
            #     # Handle the case where no text was extracted from both methods
            print("error")
            return render(request, 'analysis.html', {'error': f'Error extracting text: {str(e)}'})
            

        # Now, send the extracted text to OpenAI to extract structured data
        try:
            ai_response = process_pdf_with_gpt(text)
        except Exception as e:
            # Handle errors with OpenAI processing
            return render(request, 'analysis.html', {'error': f'Error processing with OpenAI: {str(e)}'})

        # Return the extracted data as JSON (or process as needed)
        return JsonResponse({"extracted_data": ai_response})

    else:
        # In case of unsupported HTTP method
        return render(request, 'analysis.html', {'error': 'Unsupported HTTP method'})
    
    import openai






# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class PDFProcessor:
    # def __init__(self):
    #     # Initialize OpenAI client
    #     # self.openai_client = openai.OpenAI(api_key=os.environ.get('OPENAI_API_KEY'))
        
    #     # Configure pytesseract path if needed
    #     # pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  # Uncomment and modify for Windows

    def extract_text_from_pdf(self, pdf_path):
        """Extract text from text-based PDFs using pdfplumber"""
        try:
            text = ""
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    extracted = page.extract_text()
                    if extracted:
                        text += extracted + "\n"
            return text
        except Exception as e:
            logger.error(f"Error extracting text with pdfplumber: {str(e)}")
            return ""

    def extract_text_from_image(self, image):
        """Extract text from a single image using Tesseract"""
        try:
            # Improve image quality for OCR
            image = image.convert('L')  # Convert to grayscale
            return pytesseract.image_to_string(image, lang='eng')
        except Exception as e:
            logger.error(f"Error extracting text from image: {str(e)}")
            return ""

    def extract_text_from_scanned_pdf(self, pdf_path):
        """Extract text from scanned PDFs using pdf2image and Tesseract"""
        try:
            text = ""
            # Create temporary directory for images
            with tempfile.TemporaryDirectory() as temp_dir:
                # Convert PDF to images
                images = convert_from_path(
                    pdf_path,
                    output_folder=temp_dir,
                    fmt='png',
                    dpi=300  # Higher DPI for better quality
                )
                
                # Process each page
                for i, image in enumerate(images):
                    page_text = self.extract_text_from_image(image)
                    if page_text:
                        text += f"\n--- Page {i+1} ---\n{page_text}"
                
            return text
        except Exception as e:
            logger.error(f"Error processing scanned PDF: {str(e)}")
            return ""

    

def upload_pdf_for_analysis(request):
    """Handle PDF upload and processing"""
    if request.method == 'GET':
        return render(request, 'analysis.html')

    elif request.method == 'POST':
        if 'file' not in request.FILES:
            return render(request, 'analysis.html', {'error': 'No file uploaded'})

        uploaded_file = request.FILES['file']
        
        # Validate file type
        if not uploaded_file.name.lower().endswith('.pdf'):
            return render(request, 'analysis.html', {'error': 'Please upload a PDF file'})

        try:
            # Initialize processor
            processor = PDFProcessor()
            
            # Save the uploaded file
            file_path = default_storage.save(f"uploads/{uploaded_file.name}", uploaded_file)
            absolute_file_path = os.path.join(settings.MEDIA_ROOT, file_path)

            # Try text-based extraction first
            text = processor.extract_text_from_pdf(absolute_file_path)

            # If no text found, try OCR
            if not text.strip():
                logger.info("No text found with pdfplumber, attempting OCR...")
                text = processor.extract_text_from_scanned_pdf(absolute_file_path)

            if not text.strip():
                return render(request, 'analysis.html', 
                            {'error': 'Could not extract text from the PDF'})

            # Process with OpenAI
            structured_data = process_pdf_with_gpt(text)
            
            return JsonResponse({
                'status': 'success',
                'extracted_data': structured_data,
                'raw_text': text[:1000] + '...' if len(text) > 1000 else text
            })

        except Exception as e:
            logger.error(f"Error processing PDF: {str(e)}")
            return JsonResponse({
                'status': 'error',
                'error': str(e)
            })

        finally:
            # Cleanup: Remove uploaded file
            try:
                os.remove(absolute_file_path)
            except Exception as e:
                logger.error(f"Error cleaning up file: {str(e)}")



def extract_text_with_textract(pdf_path):
    """Extract text from PDFs using AWS Textract"""
    try:
        # Initialize AWS Textract client
        textract = boto3.client(
            'textract',
            aws_access_key_id=os.environ.get(''),
            aws_secret_access_key=os.environ.get(''),
            region_name=os.environ.get('AWS_REGION', 'us-east-1')
        )

        # Read the PDF file
        with open(pdf_path, 'rb') as pdf_file:
            pdf_bytes = pdf_file.read()

        # Start the Textract job
        response = textract.start_document_text_detection(
            DocumentLocation={
                'S3Object': {
                    'Bucket': os.environ.get('AWS_BUCKET_NAME'),
                    'Name': os.path.basename(pdf_path)
                }
            }
        )

        job_id = response['JobId']

        # Wait for the job to complete
        while True:
            response = textract.get_document_text_detection(JobId=job_id)
            status = response['JobStatus']
            
            if status in ['SUCCEEDED', 'FAILED']:
                break
            time.sleep(5)

        if status == 'SUCCEEDED':
            text = ""
            for item in response['Blocks']:
                if item['BlockType'] == 'LINE':
                    text += item['Text'] + "\n"
            
            # Get remaining pages if any
            next_token = response.get('NextToken')
            while next_token:
                response = textract.get_document_text_detection(
                    JobId=job_id,
                    NextToken=next_token
                )
                for item in response['Blocks']:
                    if item['BlockType'] == 'LINE':
                        text += item['Text'] + "\n"
                next_token = response.get('NextToken')
                
            return text
        else:
            raise Exception(f"Textract job failed with status: {status}")

    except ClientError as e:
        raise Exception(f"AWS Textract error: {str(e)}")

def process_pdf_with_openai(text):
    """Process extracted text with OpenAI"""
    try:
        openai.api_key = os.environ.get('OPENAI_API_KEY')
        
        prompt = f"""
        Below is the content extracted from a PDF. Please extract key details such as names, addresses, dates, and any other structured data:

        {text}

        Extracted details:
        - Name of person/entity
        - Address
        - Date (if applicable)
        - Any other relevant details
        """

        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2
        )

        return response['choices'][0]['message']['content']
        
    except Exception as e:
        raise Exception(f"OpenAI processing error: {str(e)}")

def upload_to_s3(file_path, bucket_name):
    """Upload file to S3 bucket"""
    try:
        s3 = boto3.client(
            's3',
            aws_access_key_id=os.environ.get(''),
            aws_secret_access_key=os.environ.get(''),
            region_name=os.environ.get('AWS_REGION', 'us-east-1')
        )
        
        file_name = os.path.basename(file_path)
        s3.upload_file(file_path, bucket_name, file_name)
        return True
    
    except ClientError as e:
        raise Exception(f"S3 upload error: {str(e)}")

def upload_pdf(request):
    """Handle PDF uploads and process them"""
    if request.method == 'GET':
        return render(request, 'analysis.html')

    elif request.method == 'POST':
        if 'file' not in request.FILES:
            return render(request, 'analysis.html', {'error': 'No file uploaded'})

        uploaded_file = request.FILES['file']
        
        try:
            # Save the file locally first
            file_path = default_storage.save(f"uploads/{uploaded_file.name}", uploaded_file)
            absolute_file_path = os.path.join(settings.MEDIA_ROOT, file_path)

            # Try text-based extraction first
            text = extract_text_from_pdf(absolute_file_path)

            # If no text extracted, use AWS Textract
            if not text.strip():
                # Upload to S3 first (Textract requirement)
                upload_to_s3(absolute_file_path, os.environ.get('AWS_BUCKET_NAME'))
                text = extract_text_with_textract(absolute_file_path)

            # Process with OpenAI
            if text:
                ai_response = process_pdf_with_openai(text)
                return JsonResponse({"extracted_data": ai_response})
            else:
                return render(request, 'analysis.html', 
                            {'error': 'No text could be extracted from the PDF'})

        except Exception as e:
            return render(request, 'analysis.html', {'error': str(e)})
        
        finally:
            # Cleanup: Remove local file
            if os.path.exists(absolute_file_path):
                os.remove(absolute_file_path)

    else:
        return render(request, 'analysis.html', 
                     {'error': 'Unsupported HTTP method'})

def play_with_gemini(request):
    endpoint = "https://api.gemini-flash.com/v1/extract/bank-statement"
    api_key =  os.getenv("GEMINI_API_KEY")
    if request.method == 'GET':
        return render(request, 'analysis.html')

    elif request.method == 'POST':
        if 'file' not in request.FILES:
            return render(request, 'analysis.html', {'error': 'No file uploaded'})

        uploaded_file = request.FILES['file']
        
        # Validate file type
        if not uploaded_file.name.lower().endswith('.pdf'):
            return render(request, 'analysis.html', {'error': 'Please upload a PDF file'})
        
        file_path = default_storage.save(f"uploads/{uploaded_file.name}", uploaded_file)
        absolute_file_path = os.path.join(settings.MEDIA_ROOT, file_path)
        pdf_file_name = uploaded_file.name


        with open(absolute_file_path, "rb") as f:
            pdf_data = base64.b64encode(f.read())

        # model for how the response should be
        extract_config = {
            "financialRecords": {
                "transactionTypes": ["debit", "credit"],
                "fields": ["date", "description", "amount", "balance"]
            }   
        }
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        data = {
            "file": {
                "data": pdf_data.decode("utf-8"),
                "filename": pdf_file_name,
                "contentType": "application/pdf"
            },
            "extractConfig": extract_config
        }

        response = requests.post(endpoint, headers=headers, json=data)

        # Check response status code
        if response.status_code == 200:
            # Parse response JSON
            response_data = response.json()
            print(json.dumps(response_data, indent=4))
        else:
            print(f"Error: {response.status_code}")

def play_with_gemini_using_gemini_client(request):
    from google import genai
    from google.genai import types
    import pathlib
    import httpx

    if request.method == 'GET':
        return render(request, 'analysis.html')

    elif request.method == 'POST':
        if 'file' not in request.FILES:
            return render(request, 'analysis.html', {'error': 'No file uploaded'})

        uploaded_file = request.FILES['file']

        client = genai.Client(api_key='AIzaSyABOaMZVSeK4aDA_-qEBuVUCYPixoviZbQ')

        uploaded_file = request.FILES['file']

        # Retrieve and encode the PDF byte
        file_path = default_storage.save(f"uploads/{uploaded_file.name}", uploaded_file)
        absolute_file_path = os.path.join(settings.MEDIA_ROOT, file_path)
        pdf_file_name = uploaded_file.name

        extract_config = {
            "financialRecords": {
                "transactionTypes": ["debit", "credit"],
                "fields": ["date", "description", "amount", "balance"]
            }   
        }

        with open(absolute_file_path, "rb") as f:
            pdf_data = base64.b64encode(f.read())


        prompt = f"Accurately extract all the transaction data made on the account in from all the pages of this document and arrange each transaction into {extract_config} format."
        response = client.models.generate_content(
        model="gemini-2.0-flash",
        contents=[
            types.Part.from_bytes(
                data= pdf_data.decode("utf-8"),
                mime_type='application/pdf',
            ),
            prompt])
        print(response.text)
        return JsonResponse({
            'status': 'success',
            'extracted_data': response.text,
        })

from google import genai
from google.genai import types
import pathlib
import httpx

def play_with_gemini_using_gemini_client_for_large(request):
    from google.genai.types import HttpOptions
    import pathlib
    import httpx
    
    if request.method == 'GET':
        return render(request, 'analysis.html')

    elif request.method == 'POST':
        GEMINI_TIMEOUT = 10 * 60 * 1000  #  minutes        
        client = genai.Client(api_key='AIzaSyABOaMZVSeK4aDA_-qEBuVUCYPixoviZbQ', http_options=types.HttpOptions(timeout=600000))

        uploaded_file = request.FILES['file']
        
        try:
            print('analyzing uploaded')
            
            # Retrieve and encode the PDF byte
            file_path = default_storage.save(f"uploads/{uploaded_file.name}", uploaded_file)
            absolute_file_path = os.path.join(settings.MEDIA_ROOT, file_path)
            pdf_file_name = uploaded_file.name

        
            # Upload the PDF using the File API
            sample_file = client.files.upload(
                file=absolute_file_path,
            )

            extract_config = {                
                "transactionTypes": ["debit", "credit"],
                "fields": ["date", "description", "amount",]                  
            }

            prompt = f"Accurately extract and arragne all the transaction record from all the pages of this document and arrange each transaction into the following format: {extract_config}"

            response = client.models.generate_content(
            model="gemini-2.0-flash",
            contents=[sample_file, prompt])
            print(response.text)

            # Return the response text
            return render(request, 'analysis.html', {'extracted_data': response.text})

        except Exception as e:
            print('error', e)
        finally:
             # Cleanup: Remove local file
            if os.path.exists(absolute_file_path):
                os.remove(absolute_file_path)


        # Retrieve and encode the PDF byte
        file_path = default_storage.save(f"uploads/{uploaded_file.name}", uploaded_file)
        absolute_file_path = os.path.join(settings.MEDIA_ROOT, file_path)
        pdf_file_name = uploaded_file.name

        # file_path = pathlib.Path(pdf_file_name)
        # file_path.write_bytes(httpx.get(absolute_file_path).content)

       
        # Upload the PDF using the File API
        sample_file = client.files.upload(
            file=absolute_file_path,
        )

        extract_config = {
            "financialRecords": {
                # "transactionTypes": ["debit", "credit"],
                "fields": ["date", "description", "amount",]
            }   
        }

        prompt = f"Accurately extract all the transaction data from all the pages of this document and arrange each transaction into the following format: {extract_config}"

        response = client.models.generate_content(
        model="gemini-2.0-flash",
        contents=[sample_file, prompt])
        print(response.text)

        # Return the response text
        return render(request, 'analysis.html', {'extracted_data': response.text})
    
def transcribe_call_with_gemini(request):
    from google import genai
    google_api_key ='...'
    api_key = google_api_key
    client = genai.Client(api_key=google_api_key)

    if request.method == 'GET':
        return render(request, 'call_audit.html')
    elif request.method == 'POST':
        
        try:
            # Check if the request has a file
            if 'audio_file' not in request.FILES:
                return JsonResponse({'error': 'No audio file provided'}, status=400)
            
            audio_file = request.FILES['audio_file']
            disposition = request.POST.get('disposition')
            agent_id = request.POST.get('agent_id')
            agent_name = request.POST.get('agent_name')
            auditor = "Rilwan Hassan"
            
            # Gemini API key
            # api_key = os.environ.get('GEMINI_API_KEY')
            
            if not api_key:
                return JsonResponse({'error': 'Gemini API key not provided'}, status=400)
                        
            
            # Retrieve and encode the PDF byte
            file_path = default_storage.save(f"uploads/{audio_file.name}", audio_file)
            absolute_file_path = os.path.join(settings.MEDIA_ROOT, file_path)
            pdf_file_name = audio_file.name

        
            # Upload the PDF using the File API
            sample_file = client.files.upload(
                file=absolute_file_path,
            )
            
            try:

                audit_guidelines = """
                    Use the following company-defined criteria to audit this collection call:

                    1. Foundation (15%): if the agent says or attempts to introduce themself by sayying their name, where and why they are calling and the customer name and indicacte that the call is being recorded give them full marks, but if the agent does not notify the customer that the call is being recorded give them 0 marks even if the did the other part of the intoduction.
                    2. Communication Skills (25%): Assess clarity, toneand firmness and use of appropriate language always give them full marks here.
                    3. Probing Skills (20%): Check how effectively the agent asked open-ended questions to understand customer needs if the customer mentions that hte product is faulty or payment related issue.
                    4. Negotiation Skills (35%): The agent should always mention payment amount and date of payment if {disposition} is Customer reached sucessfully, but if {disposition} is Promise to pay either payment date or payent amount should be mention or at least a promise to pay by the customer.
                    5. Closing Skills (5%): The agent should reiterate commintments and timelines and did the agent do closing verbage.

                    The {disposition} is the final outcome of the call. If the disposition does not reflect what the call was about then, it will be considered an auto-fail in Closing Skills.
                    Auto Fail means 0 in that parameter

                    if the {disposition} is Customer reached sucessfully or Promised to pay then all parameters needs to be checked unless the customer hung up before the agent could finish then give them full score in the emaining parameters that they could not do before the customer hung up, but if the {disposition} is soething else then Negotiation Skills is no compulsory so give them full marks there and ensure that the disposition reflects the sumaary of the call or must at least be part of the themes in that call.



                    Return the audit result in JSON format with the scores and short justifications per category and a total score.
                    """

                audit_guides = """
                    Use the following company-defined criteria to audit this collection call. Assess each parameter and provide a JSON result with scores and short justifications. Each parameter is weighted.

                    1. Foundation (15%):
                    - Full marks if the agent introduces themselves by stating their name, the company, the purpose of the call, confirms the customer name, and informs the customer that the call is being recorded.
                    - If the agent fails to inform the customer that the call is being recorded, they must receive **0 marks** for this section, even if the other parts are done correctly.
                    - If the agent does not mention their name or the customer name or the company name or the purpose of the call, they should receive **10 marks** for this section.

                    2. Communication Skills (25%):
                    - Assess clarity, tone, firmness, and use of appropriate language.
                    - Assume agents should always receive full marks unless they demonstrate poor communication.

                    3. Probing Skills (20%):
                    - If the customer mentions a faulty product or payment-channel-related issue, check how well the agent asked open-ended questions to understand the customer's needs.

                    4. Negotiation Skills (35%):
                    - If the {disposition} is **"Customer reached successfully"**, the agent must mention both the **payment amount** and **payment date** to get full marks.
                    - If the {disposition} is **"Promise to Pay"**, the agent must mention **either** the payment date, payment amount, a **promise to pay** from the customer or at payment must have been discussed by either parties on the call.
                    - if the {disposition} is **Already Enabled/Unlocked** it should be established that the customer has made payment before the call was made.
                    - Missing these details should result in reduced or zero marks.

                    5. Closing Skills (5%):
                    - The agent should reiterate the customer's commitments or timelines, and use proper closing language (closing verbiage).
                    - If the {disposition} does not reflect the actual conversation or is clearly wrong, mark **0** for this section (Auto-fail).

                    Additional Rules:
                    - **Auto Fail** means 0 for that parameter.
                    - If the {disposition} is **"Customer reached successfully"** or **"Promised to pay"**, evaluate all parameters unless the customer **hung up prematurely**. If so, award full marks for any remaining sections the agent couldn't complete.
                    - If the {disposition} is something else, then **Negotiation Skills is not compulsory**—award full marks for it—but ensure the disposition reflects the overall theme or summary of the call.
                    -if the {disposition} is **"No Disposition Selected"** or **"No Disposition"**, the agent should receive **0 marks** for Closing Skills parameter.

                    Return the audit result in JSON format, structured with:
                    - Individual parameter scores and justifications
                    - A total score
                    - Language of the call
                    """
                             
                prompt_1 = f"""
                    As a collection call quality assurance agent, accurately audit the call using the criteria and weights described.
                    {audit_guides}
                    """
                                
                response2 = client.models.generate_content(
                    model="gemini-2.0-flash",
                    contents=[sample_file, prompt_1])

                
                raw_text = response2.candidates[0].content.parts[0].text
                match = re.search(r"```json\s*(\{.*?\})\s*```", raw_text, re.DOTALL)
                if match:
                    cleaned_json = match.group(1)
                else:
                    # Fallback: try to extract first valid-looking JSON object manually
                    start = raw_text.find('{')
                    end = raw_text.rfind('}') + 1
                    cleaned_json = raw_text[start:end]
                try:
                    parsed_json = json.loads(cleaned_json)
                    
                    audit_result = parsed_json.get("audit_results", {})
                    

                    # save the audit
                    try:
                        new_audit = CallAudit(
                        name=agent_name,
                        auidited_by=auditor,
                        agent_id=agent_id,
                        agent_disposition=disposition,
                        call_language=parsed_json.get("language", "English"),
                        foundation_Skill_score=audit_result.get("Foundation", {}).get("score", 0),
                        commuication_skill_score=audit_result.get("Communication Skills", {}).get("score", 0),
                        probing_skill_score=audit_result.get("Probing Skills", {}).get("score", 0),
                        negotiation_skill_score=audit_result.get("Negotiation Skills", {}).get("score", 0),
                        closing_skill_score=audit_result.get("Closing Skills", {}).get("score", 0),
                        total_score=parsed_json.get("total_score", 0),
                        # justification=audit_result  
                        )
                        new_audit.save()
                        
                    except Exception as e:
                        print('error', e)
                        return JsonResponse({
                            'success': False,
                            'error': 'Failed to save audit result',
                            'details': str(e),
                        })
                    
                except json.JSONDecodeError as e:
                    return JsonResponse({
                        'success': False,
                        'error': 'Failed to parse audit result',
                        'details': str(e),
                        'raw_response': raw_text  # helpful for debugging
                    })
                
                
                return JsonResponse({
                    'success': True,
                    'audit': parsed_json,
                   
                })
                    
            finally:
                # Clean up temporary file
                if os.path.exists(file_path):
                    os.unlink(file_path)
        
        except Exception as e:
            logger.error(f"Error in transcribe_call: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

def get_mime_type(filename):
    """
    Determine the MIME type based on the file extension.
    """
    extension = os.path.splitext(filename)[1].lower()
    mime_types = {
        '.mp3': 'audio/mp3',
        '.mp4': 'audio/mp4',
        '.m4a': 'audio/m4a',
        '.wav': 'audio/wav',
        '.flac': 'audio/flac',
        '.ogg': 'audio/ogg',
    }
    return mime_types.get(extension, 'audio/mpeg')

def get_all_audit(request):
    if request.method == 'GET':
        audits = CallAudit.objects.all().order_by('-created_at')
        return render(request, 'call_audit_table_page.html', {'audits': audits})
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)
    
from django.shortcuts import get_object_or_404

def get_single_audit(request, uid):
    if request.method == 'GET':
        audit = get_object_or_404(CallAudit, uid=uid)
        return render(request, 'call_audit_detail_page.html', {'audit': audit})
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)

def call_call_overdue_customers(request):
    if request.method == 'GET':
        print('calling overdue customers')
    elif request.method == 'POST':
        from twilio.rest import Client 
        account_sid = os.getenv("OPENAI_API_KEY")
        auth_token = os.getenv("OPENAI_API_KEY")
        client = Client(account_sid, auth_token)

        call = client.calls.create(
            url="http://demo.twilio.com/docs/voice.xml",
            to="2349074452956",
            from_="+17167064540",
        )

        print(call.sid)
        print('calling overdue customers')


import openpyxl
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from twilio.rest import Client


def call_call_overdue_lamp_customers(request):
    if request.method == 'GET':
        return render(request, 'call_customers.html')

    elif request.method == 'POST':
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file uploaded'}, status=400)

        excel_file = request.FILES['file']

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active  
        except Exception as e:
            return JsonResponse({'error': f'Invalid Excel file: {str(e)}'}, status=400)

        account_sid = os.getenv("OPENAI_API_KEY")
        auth_token = os.getenv("OPENAI_API_KEY")
        client = Client(account_sid, auth_token)

        calls_made = []

        # Skip header row 
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, amount_due, phone_number = row

            # Convert to E.164 format if necessary
            to_number = str(phone_number)
            if not to_number.startswith('+') and not to_number.startswith('234'):
                to_number = '234' + to_number.lstrip('0')

            try:
                call = client.calls.create(
                    url=f"https://4c4d-105-113-107-46.ngrok-free.app/prospects/twiml-message/?name={name}&amount={amount_due}", #message
                    to='=2349074452956',
                    from_="+17167064540",
                )
                calls_made.append({'name': name, 'phone': to_number, 'sid': call.sid})
            except Exception as e:
                calls_made.append({'name': name, 'phone': to_number, 'error': str(e)})

        return JsonResponse({'calls': calls_made})

from django.views.decorators.csrf import csrf_exempt
import xml.etree.ElementTree as ET

@csrf_exempt
def twiml_message(request):
    name = request.GET.get('name', 'Customer')
    amount = request.GET.get('amount', 'an outstanding balance')

    # Build TwiML XML response
    response = ET.Element('Response')
    say = ET.SubElement(response, 'Say', voice="alice")
    say.text = f"Hello {name}, you have an outstanding balance of {amount} Naira. Please make your payment as soon as possible."

    xml_string = ET.tostring(response, encoding='utf-8', method='xml')
    return HttpResponse(xml_string, content_type='text/xml')




@csrf_exempt
def call_call_overdue_lamp_customers_with_twilio(request):
    if request.method == 'GET':
        return render(request, 'call_customers.html')

    elif request.method == 'POST':
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file uploaded'}, status=400)

        excel_file = request.FILES['file']

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active  
        except Exception as e:
            return JsonResponse({'error': f'Invalid Excel file: {str(e)}'}, status=400)

        account_sid = os.getenv("OPENAI_API_KEY")
        auth_token = os.getenv("OPENAI_API_KEY")
        client = Client(account_sid, auth_token)

        calls_made = []

        # Skip header row 
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, amount_due, phone_number = row

            # Convert to E.164 format if necessary
            # request.user.country_code
            country_code = '234' 
            to_number = str(phone_number)
            if not to_number.startswith('+') and not to_number.startswith('234'):
                to_number = '234' + to_number.lstrip('0')

            try:
                name = "abisola"
                amount_due = 15000
                # url=f"https://4c4d-105-113-107-46.ngrok-free.app/prospects/twiml-message/?name={name}&amount={amount_due}", #message
                call = client.calls.create(
                    twiml=f"<Response><Say voice='alice'>Hello {name}, you have an outstanding balance of {amount_due} Naira. Please make your payment as soon as possible.</Say></Response>",                    
                    to="+2349074452956",
                    from_="+17167064540",
                )
                calls_made.append({'name': name, 'phone': to_number, 'sid': call.sid})
            except Exception as e:
                calls_made.append({'name': name, 'phone': to_number, 'error': str(e)})

        return JsonResponse({'calls': calls_made})
    


import openpyxl
from django.shortcuts import render

from django.conf import settings
from twilio.rest import Client
from twilio.base.exceptions import TwilioException


def call_overdue_lamp_customers(request):
    if request.method == 'GET':
        return render(request, 'call_customers.html')

    elif request.method == 'POST':
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file uploaded'}, status=400)

        excel_file = request.FILES['file']

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active  
        except Exception as e:
            logger.error(f"Excel file loading error: {str(e)}")
            return JsonResponse({'error': f'Invalid Excel file: {str(e)}'}, status=400)

        # Move credentials to settings for security
        account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', os.getenv("OPENAI_API_KEY"))
        auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', os.getenv("OPENAI_API_KEY"))
        from_number = getattr(settings, 'TWILIO_FROM_NUMBER', '+17167064540')
        
        client = Client(account_sid, auth_token)
        calls_made = []

        # Skip header row 
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Handle empty rows
            if not row or all(cell is None for cell in row):
                continue
                
            # Ensure we have at least 3 columns
            if len(row) < 3:
                calls_made.append({'error': 'Insufficient data in row', 'row_data': str(row)})
                continue
                
            name, amount_due, phone_number = row[:3]  # Take first 3 columns
            
            # Validate required fields
            if not name or not amount_due or not phone_number:
                calls_made.append({
                    'name': name or 'Unknown',
                    'error': 'Missing required data (name, amount_due, or phone_number)'
                })
                continue

            # Format phone number to E.164 format
            country_code = '234'  # Nigeria country code
            to_number = str(phone_number).strip()
            
            # Remove any non-numeric characters except +
            import re
            to_number = re.sub(r'[^\d+]', '', to_number)
            
            if not to_number.startswith('+'):
                if to_number.startswith('234'):
                    to_number = '+' + to_number
                elif to_number.startswith('0'):
                    to_number = '+234' + to_number[1:]
                else:
                    to_number = '+234' + to_number

            try:
                # Validate phone number format
                if len(to_number) < 10 or not to_number.startswith('+234'):
                    raise ValueError(f"Invalid phone number format: {to_number}")
                
                # Format amount for speech
                amount_formatted = f"{amount_due:,.0f}" if isinstance(amount_due, (int, float)) else str(amount_due)
                
                # Create TwiML with proper escaping
                twiml_message = f"""
                <Response>
                    <Say voice='alice'>
                        Hello {str(name).replace('&', 'and').replace('<', '').replace('>', '')}, 
                        you have an outstanding balance of {amount_formatted} Naira. 
                        Please make your payment as soon as possible. Thank you.
                    </Say>
                </Response>
                """.strip()

                call = client.calls.create(
                    twiml=twiml_message,
                    to=to_number,
                    from_=from_number,
                )
                
                calls_made.append({
                    'name': str(name),
                    'phone': to_number,
                    'amount_due': amount_formatted,
                    'sid': call.sid,
                    'status': 'queued'
                })
                
                logger.info(f"Call initiated for {name} at {to_number}, SID: {call.sid}")
                
            except TwilioException as e:
                error_msg = f"Twilio error: {str(e)}"
                calls_made.append({
                    'name': str(name),
                    'phone': to_number,
                    'amount_due': str(amount_due),
                    'error': error_msg
                })
                logger.error(f"Twilio error for {name}: {error_msg}")
                
            except Exception as e:
                error_msg = f"General error: {str(e)}"
                calls_made.append({
                    'name': str(name),
                    'phone': to_number,
                    'amount_due': str(amount_due),
                    'error': error_msg
                })
                logger.error(f"Error processing call for {name}: {error_msg}")

        # Summary statistics
        successful_calls = [call for call in calls_made if 'sid' in call]
        failed_calls = [call for call in calls_made if 'error' in call]
        
        return JsonResponse({
            'calls': calls_made,
            'summary': {
                'total_processed': len(calls_made),
                'successful': len(successful_calls),
                'failed': len(failed_calls)
            }
        })

    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)


import africastalking
import pandas as pd
from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt


# Set up logging
logger = logging.getLogger(__name__)

# Initialize Africa's Talking
username = "RobocallApP"  
api_key = os.getenv("AFRICAS_TALKING_API_KEY") 
africastalking.initialize(username, api_key)
voice = africastalking.Voice

# Your registered test or business number
CALL_FROM_NUMBER = "+2342017001133"
VOICE_CALLBACK_BASE_URL = "https://b09ff8012dbb.ngrok-free.app/prospects/voice/response/"  

@csrf_exempt
def call_overdue_customers_with_ait(request):
    if request.method == 'GET':
        return render(request, 'call_customers.html')

    elif request.method == 'POST':
        print("Received POST request to call overdue customers")
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file uploaded'}, status=400)

        excel_file = request.FILES['file']

        print("Received POST request  file to call overdue customers")

        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            logger.error(f"Failed to read Excel file: {str(e)}")
            return JsonResponse({'error': f'Invalid Excel file: {str(e)}'}, status=400)
        
        print("read file  to call overdue customers")

        # required_columns = {"name", "phone", "amount_due"}
        # if not required_columns.issubset(df.columns):
        #     print("Failed to get details retrieved to call overdue customers")
        #     return JsonResponse({'error': f'Missing required columns: {required_columns}'}, status=400)
        
        print("Details retrieved to call overdue customers")

        call_results = []
        for index, row in df.iterrows():
            name = "Rilwan"
            phone = "+2347037971851"  # Ensure phone number is in E.164 format 
            

            try:
                response = voice.call(
                    callTo = [phone],
                    callFrom = CALL_FROM_NUMBER,
                    # url=callback_url
                )
                call_results.append({'name': name, 'phone': phone, 'status': 'success', 'response': response})
            except Exception as e:
                logger.error(f"Call failed for {phone}: {str(e)}")
                call_results.append({'name': name, 'phone': phone, 'status': 'error', 'error': str(e)})

        print("Call results:", call_results)

        return JsonResponse({'results': call_results})


# @csrf_exempt
# def ait_voice_response(request):
#     xml = """<?xml version="1.0" encoding="UTF-8"?>
#         <Response>
#             <Say voice="en-US-Standard-C" playBeep="false" >Dear customer, This is a reminder from SunKing Solar that your product is due for payment today</Say>
#         </Response>
#     """
#     return HttpResponse(xml, content_type="application/xml")



@csrf_exempt
def ait_voice_response(request):
    xml = """<?xml version="1.0" encoding="UTF-8"?>
        <Response>
            <Say voice="en-US-Standard-C" playBeep="false">Dear customer, This is a reminder from SunKing Solar.Your SunKing product is due for payment. Kindly make your payment today. Thank you.
            </Say>
        </Response>
    """
    return HttpResponse(xml, content_type="application/xml")


import africastalking
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
import logging
import os
from datetime import datetime, timedelta
import time
import json
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

# Set up logging
logger = logging.getLogger(__name__)

# Initialize Africa's Talking
username = os.getenv("AFRICASTALKING_USERNAME"  "RobocallApP") 
api_key = os.getenv("AFRICAS_TALKING_API_KEY") 
africastalking.initialize(username, api_key)
voice = africastalking.Voice

# Your registered test or business number
CALL_FROM_NUMBER = "+2342017001133"
VOICE_CALLBACK_BASE_URL = "https://b09ff8012dbb.ngrok-free.app/prospects/voice/response/"

# Google Sheets setup
SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']

# Rate limiting settings (based on Africa's Talking guidelines)
MAX_CALLS_PER_MINUTE = 30  # Conservative estimate
CALL_DURATION_ESTIMATE = 45  # Average call duration in seconds

# Global variables for tracking call progress
call_progress = {
    'total': 0,
    'completed': 0,
    'successful': 0,
    'failed': 0,
    'start_time': None,
    'estimated_completion': None,
    'current_batch': 0,
    'total_batches': 0
}
progress_lock = threading.Lock()

def get_google_sheet_data(sheet_url, worksheet_name=None, credentials_path=None):
    """
    Fetch data from Google Sheets
    """
    try:
        if credentials_path:
            creds = Credentials.from_service_account_file(credentials_path, scopes=SCOPES)
        else:
            # Use environment variable or default service account
            creds = Credentials.from_service_account_info(
                json.loads(os.getenv('GOOGLE_SERVICE_ACCOUNT_JSON')),
                scopes=SCOPES
            )
        
        client = gspread.authorize(creds)
        
        # Open the spreadsheet
        spreadsheet = client.open_by_url(sheet_url)
        
        # Select worksheet
        if worksheet_name:
            worksheet = spreadsheet.worksheet(worksheet_name)
        else:
            worksheet = spreadsheet.sheet1
        
        # Get all data
        data = worksheet.get_all_records()
        return pd.DataFrame(data)
        
    except Exception as e:
        logger.error(f"Failed to fetch Google Sheet data: {str(e)}")
        raise

def validate_phone_number(phone):
    """
    Validate and format phone number for Africa's Talking
    """
    # Remove any non-digit characters
    phone = ''.join(filter(str.isdigit, str(phone)))
    
    # Handle Nigerian numbers (add country code if missing)
    if phone.startswith('0') and len(phone) == 11:
        phone = '+234' + phone[1:]
    elif not phone.startswith('+') and len(phone) == 10:
        phone = '+234' + phone
    
    return phone

def make_single_call(phone, name, amount_due, retry_count=0):
    """
    Make a single call with retry logic
    """
    max_retries = 2
    
    try:
        response = voice.call(
            callTo=[phone],
            callFrom=CALL_FROM_NUMBER,
            # url=callback_url  # Uncomment if you want to use callback
        )
        
        return {
            'name': name, 
            'phone': phone, 
            'amount_due': amount_due,
            'status': 'success', 
            'response': response,
            'timestamp': datetime.now().isoformat(),
            'retries': retry_count
        }
        
    except Exception as e:
        if retry_count < max_retries:
            # Wait before retrying
            time.sleep(2)
            return make_single_call(phone, name, amount_due, retry_count + 1)
        else:
            logger.error(f"Call failed for {phone} after {max_retries} retries: {str(e)}")
            return {
                'name': name, 
                'phone': phone, 
                'amount_due': amount_due,
                'status': 'error', 
                'error': str(e),
                'timestamp': datetime.now().isoformat(),
                'retries': retry_count
            }

def process_batch(batch_df, batch_number, total_batches):
    """
    Process a batch of calls with proper rate limiting
    """
    batch_results = []
    
    with progress_lock:
        call_progress['current_batch'] = batch_number
    
    for index, row in batch_df.iterrows():
        name = row.get('name', 'Customer')
        raw_phone = row.get('phone')
        amount_due = row.get('amount_due', 0)
        
        if pd.isna(raw_phone) or not raw_phone:
            result = {
                'name': name, 
                'phone': 'Missing', 
                'status': 'error', 
                'error': 'Phone number missing',
                'timestamp': datetime.now().isoformat()
            }
            batch_results.append(result)
            
            with progress_lock:
                call_progress['completed'] += 1
                call_progress['failed'] += 1
            continue
        
        # Format phone number
        phone = validate_phone_number(raw_phone)
        
        # Make the call
        result = make_single_call(phone, name, amount_due)
        batch_results.append(result)
        
        # Update progress
        with progress_lock:
            call_progress['completed'] += 1
            if result['status'] == 'success':
                call_progress['successful'] += 1
            else:
                call_progress['failed'] += 1
            
            # Update estimated completion time
            elapsed = (datetime.now() - call_progress['start_time']).total_seconds()
            calls_per_second = call_progress['completed'] / elapsed if elapsed > 0 else 0
            if calls_per_second > 0:
                remaining = call_progress['total'] - call_progress['completed']
                eta_seconds = remaining / calls_per_second
                call_progress['estimated_completion'] = (
                    datetime.now() + timedelta(seconds=eta_seconds)
                ).isoformat()
        
        # Rate limiting - sleep between calls
        time.sleep(60 / MAX_CALLS_PER_MINUTE)
    
    return batch_results

@csrf_exempt
def call_overdue_customers_with_ait(request):
    if request.method == 'GET':
        return render(request, 'call_customers.html')

    elif request.method == 'POST':
        print("Received POST request to call overdue customers")
        
        # Reset progress
        with progress_lock:
            call_progress.update({
                'total': 0,
                'completed': 0,
                'successful': 0,
                'failed': 0,
                'start_time': datetime.now(),
                'estimated_completion': None,
                'current_batch': 0,
                'total_batches': 0
            })
        
        source_type = request.POST.get('source_type', 'file')
        batch_size = int(request.POST.get('batch_size', 100))
        max_workers = int(request.POST.get('max_workers', 1))
        
        try:
            if source_type == 'file':
                if 'file' not in request.FILES:
                    return JsonResponse({'error': 'No file uploaded'}, status=400)

                excel_file = request.FILES['file']
                print("Processing uploaded Excel file")
                
                try:
                    df = pd.read_excel(excel_file)
                except Exception as e:
                    logger.error(f"Failed to read Excel file: {str(e)}")
                    return JsonResponse({'error': f'Invalid Excel file: {str(e)}'}, status=400)
            
            elif source_type == 'google_sheets':
                sheet_url = request.POST.get('sheet_url')
                worksheet_name = request.POST.get('worksheet_name')
                
                if not sheet_url:
                    return JsonResponse({'error': 'Google Sheet URL is required'}, status=400)
                
                print(f"Fetching data from Google Sheet: {sheet_url}")
                df = get_google_sheet_data(sheet_url, worksheet_name)
            
            else:
                return JsonResponse({'error': 'Invalid source type'}, status=400)
            
            print(f"Retrieved {len(df)} records from {source_type}")
            
            # Validate required columns
            required_columns = {"name", "phone", "amount_due"}
            if not required_columns.issubset(df.columns):
                missing = required_columns - set(df.columns)
                return JsonResponse({'error': f'Missing required columns: {missing}'}, status=400)
            
            # Update total count
            with progress_lock:
                call_progress['total'] = len(df)
            
            # Process in batches
            total_batches = (len(df) + batch_size - 1) // batch_size
            
            with progress_lock:
                call_progress['total_batches'] = total_batches
            
            all_results = []
            
            # Use ThreadPoolExecutor for parallel processing if allowed
            if max_workers > 1:
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    # Split into batches
                    futures = []
                    for i in range(total_batches):
                        start_idx = i * batch_size
                        end_idx = min((i + 1) * batch_size, len(df))
                        batch_df = df.iloc[start_idx:end_idx]
                        
                        futures.append(
                            executor.submit(process_batch, batch_df, i+1, total_batches)
                        )
                    
                    for future in as_completed(futures):
                        batch_results = future.result()
                        all_results.extend(batch_results)
            else:
                # Sequential processing
                for i in range(total_batches):
                    start_idx = i * batch_size
                    end_idx = min((i + 1) * batch_size, len(df))
                    batch_df = df.iloc[start_idx:end_idx]
                    
                    batch_results = process_batch(batch_df, i+1, total_batches)
                    all_results.extend(batch_results)
            
            # Calculate final statistics
            total_time = (datetime.now() - call_progress['start_time']).total_seconds()
            calls_per_minute = (call_progress['completed'] / total_time) * 60 if total_time > 0 else 0
            
            print(f"Calling completed: {call_progress['successful']} successful, {call_progress['failed']} failed")
            
            return JsonResponse({
                'results': all_results,
                'summary': {
                    'total_calls': call_progress['total'],
                    'successful_calls': call_progress['successful'],
                    'failed_calls': call_progress['failed'],
                    'total_time_seconds': total_time,
                    'calls_per_minute': calls_per_minute,
                    'completion_time': datetime.now().isoformat()
                }
            })
            
        except Exception as e:
            logger.error(f"Unexpected error in call processing: {str(e)}")
            return JsonResponse({'error': f'Processing failed: {str(e)}'}, status=500)

@csrf_exempt
def get_call_progress(request):
    """
    Endpoint to check progress of ongoing call campaign
    """
    with progress_lock:
        progress_data = call_progress.copy()
    
    if progress_data['completed'] > 0 and progress_data['total'] > 0:
        progress_percent = (progress_data['completed'] / progress_data['total']) * 100
    else:
        progress_percent = 0
        
    return JsonResponse({
        'progress': progress_percent,
        'details': progress_data
    })

@csrf_exempt
def ait_voice_response(request):
    """
    Voice response handler for Africa's Talking
    """
    xml = """<?xml version="1.0" encoding="UTF-8"?>
        <Response>
            <Say voice="en-US-Standard-C" playBeep="false">
                Dear customer, This is a reminder from SunKing Solar. 
                Your SunKing product is due for payment. 
                Kindly make your payment today. Thank you.
            </Say>
        </Response>
    """
    return HttpResponse(xml, content_type="application/xml")









































































# def call_overdue_customers_with_sendChamp(request):
#     if request.method == 'GET':
#         return render(request, 'call_customers.html')
#     elif request.method == 'POST':
        
#         name = "abisola"
#         amount_due = 15000

#         # Senchamp API setup
#         SENCHAMP_API_KEY = 'your_senchamp_api_key'  # <-- replace this with your real key
#         SENCHAMP_CALL_URL = 'https://api.senchamp.com/v1/voice/call'  # <-- adjust if your endpoint differs

#         # Build message
#         message = f"Hello {name}, you have an outstanding balance of {amount_due} Naira. Please make your payment as soon as possible."

#         # Send call
#         headers = {
#             'Authorization': f'Bearer {SENCHAMP_API_KEY}',
#             'Content-Type': 'application/json',
#         }

#         payload = {
#             'to': '+2349074452956',
#             'from': '+17167064540',  # your Senchamp verified number
#             'message': message,
            
#         }

#         response = requests.post(SENCHAMP_CALL_URL, json=payload, headers=headers)

#         if response.status_code == 200:
#             print("Call placed successfully:", response.json())
#         else:
#             print("Failed to place call:", response.status_code, response.text)


